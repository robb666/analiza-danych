import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
from matplotlib.dates import MonthLocator
import re


desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')


def odczyt_excela():
    wb = load_workbook(desktop + "\\2014 BAZA MAGRO.xlsx", read_only=False, data_only=True)
    ws = wb['BAZA 2014']
    OWCA = ws['G']
    nazwisko_col = ws['L']
    przypis = ws['AV']
    rozl = ws['BY']

    id_val = []
    przypis_val = []

    for i in range(len(OWCA)):
        ofwca = OWCA[i].value
        nazwisko = nazwisko_col[i].value
        rozliczone = rozl[i].value
        if ofwca in ('MAGRO', 'Robert'):
            if nazwisko is not None and re.search('[0-9]', str(przypis[i].value)) \
                    and rozliczone in 'rozl':
                id_val.append(nazwisko)
                przypis_val.append(przypis[i].value)

    return wb, OWCA, nazwisko_col, id_val, przypis_val


def scalenie_danych(id_val, przypis_val):
    tpl_lista_baza = []
    for i in zip(id_val, przypis_val):
        tpl_lista_baza.append(i)

    return tpl_lista_baza


def usunięcie_duplikatów(tpl_lista_baza):
    rezultat = {}
    for k, v in tpl_lista_baza:
        rezultat[k] = rezultat.get(k, 0) + v

    return rezultat


def sortowanie(rezultat):
    sorted_d = sorted((value, key) for (key, value) in rezultat.items())
    sorted_d.reverse()
    top_60 = sorted_d[:60]
    top_120 = sorted_d[60:120]

    return top_60, top_120


def sumy_nazwiska(top_60, top_120):
    sumy_przypisu_60 = []
    nazwiska_60 = []
    sumy_przypisu_120 = []
    nazwiska_120 = []

    for i in range(len(top_60)):
        suma_60 = top_60[i][0]
        nazwisko_60 = top_60[i][1]
        sumy_przypisu_60.append(suma_60)
        nazwiska_60.append(nazwisko_60)

        suma_120 = top_120[i][0]
        nazwisko_120 = top_120[i][1]
        sumy_przypisu_120.append(suma_120)
        nazwiska_120.append(nazwisko_120)

    return sumy_przypisu_60, nazwiska_60, sumy_przypisu_120, nazwiska_120


def pokolorowanie(OWCA, nazwisko_col, nazwiska_60, nazwiska_120):
    top_60_fill = PatternFill(fgColor='A67C00', fill_type='solid')
    top_120_fill = PatternFill(fgColor='BF9B30', fill_type='solid')

    for i in range(len(nazwisko_col)):
        if OWCA[i].value in ('MAGRO', 'Robert'):
            if nazwisko_col[i].value in nazwiska_60:
                nazwisko_col[i].fill = top_60_fill
            if nazwisko_col[i].value in nazwiska_120:
                nazwisko_col[i].fill = top_120_fill

    wb.save(desktop + '\BAZA MAGRO kategorie klientów ' + '.xlsx')


def wykres_top_60(sumy_przypisu_60, nazwiska_60):

    plt.style.use('ggplot')

    fig, ax = plt.subplots()
    fig.autofmt_xdate()
    mloc = MonthLocator()
    ax.tick_params(axis='both', which='major', labelsize=8)
    ax.xaxis.set_minor_locator(mloc)
    plt.title('Klienci "Top 60" wg. przypisu w złotych od 2014 roku')
    plt.plot(nazwiska_60, sumy_przypisu_60, color='#2E8B57', label='zysk = ')
    plt.show()


def wykres_top_120(sumy_przypisu_120, nazwiska_120):

    plt.style.use('ggplot')

    fig, ax = plt.subplots()
    fig.autofmt_xdate()
    mloc = MonthLocator()
    ax.tick_params(axis='both', which='major', labelsize=8)
    ax.xaxis.set_minor_locator(mloc)
    plt.title('Klienci "Top 60 - 120" wg. przypisu w złotych od 2014 roku')
    plt.plot(nazwiska_120, sumy_przypisu_120, color='#2E8B57', label='zysk = ')
    plt.show()


wb, OWCA, nazwisko_col, id_val, przypis_val = odczyt_excela()
tpl_lista_baza = scalenie_danych(id_val, przypis_val)
rezultat = usunięcie_duplikatów(tpl_lista_baza)
top_60, top_120 = sortowanie(rezultat)
sumy_przypisu_60, nazwiska_60, sumy_przypisu_120, nazwiska_120 = sumy_nazwiska(top_60, top_120)
# pokolorowanie(OWCA, nazwisko_col, nazwiska_60, nazwiska_120)
print(wykres_top_60(sumy_przypisu_60, nazwiska_60))
print(wykres_top_120(sumy_przypisu_120, nazwiska_120))
