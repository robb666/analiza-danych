import os
from win32com.client import Dispatch
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.dates import MonthLocator

# from histogram_inkasa import inkaso_m

application_path = os.path.dirname(os.path.abspath(__file__))
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')


def sorting():
    xlApp = Dispatch("Excel.Application")
    xlwb = xlApp.Workbooks.Open(desktop + "\\2014 BAZA MAGRO.xlsm", False, False, None)
    xlAscending = 1
    xlSortColumns = 1
    xlYes = 2
    xlApp.Range("BB5:CU18500").Sort(Key1=xlApp.Range("BB5:CU18500"), Order1=xlAscending,
                                    Header=xlYes, Orientation=xlSortColumns)
    path = ''.join([desktop])
    xlApp.DisplayAlerts = False
    xlwb.SaveAs(path + '\\MODIFY.xlsm')
    xlApp.DisplayAlerts = True
    xlwb.Close()
    xlApp = None


def odczytDanych():
    wb = load_workbook(desktop + "\\MODIFY.xlsm", read_only=False, data_only=True)
    # ws = wb['Arkusz1']
    ws = wb['BAZA 2014']
    BY = ws['BY']
    CU = ws['CU']
    BB = ws['BB']

    # cells = ws['BB5':'BC16813']

    return BY, CU, BB


def określenieDanych(BY, CU, BB):
    data = []
    przychód = []
    for x in range(len(BY)):
        # if BB[x].value is None: #### ???  #####
        #     pass
        if BY[x].value == 'rozl':
            data_raw = str(BB[x].value)[:7]
            data.append(data_raw)
            prowizja = int(CU[x].value)
            przychód.append(prowizja)

    return data, przychód


def listaTupli(data, przychód):
    tupla_lista = []
    for i in zip(data, przychód):
        tupla_lista.append(i)

    return tupla_lista


def przychodyMsc(tupla_lista):
    rezultat = {}
    for k, v in tupla_lista:
        rezultat[k] = rezultat.get(k, 0) + v

    return rezultat


def selekcjaDanych(rezultat):

    czas = []
    przych = []
    for i in rezultat.keys():
        czas.append(i)
    for j in rezultat.values():
        przych.append(j)

    return czas, przych


def wykres(czas, przych):

    date_str = []
    for i in czas:
        if i != 'None':
            date_str.append(datetime.strptime(i, '%Y-%m'))
    przych.remove(195)
    print(czas)
    print(przych)
    print(date_str)

    plt.style.use('ggplot')                                              # 'fivethirtyeight'
    fig, ax = plt.subplots()
    # rotate and align the tick labels so they look better
    fig.autofmt_xdate()
    mloc = MonthLocator()
    ax.xaxis.set_minor_locator(mloc)

    # use a more precise date string for the x axis locations in the toolbar

    ax.fmt_xdata = mdates.DateFormatter('%Y-%m-%d')
    ax.set_title('fig.autofmt_xdate fixes the labels')

    plt.plot(date_str, przych, color='#008df5', label='przychody')
    # plt.plot(date_str, inkaso_m, color='#001df5', label='inkaso')
    # plt.xlabel('rok')
    plt.ylabel('Przychód od 2014 roku w zł')
    plt.title('PRZYCHÓD')
    plt.savefig(desktop + "\\przych z bazy.png")
    plt.grid(which='major', color='w', linestyle='-', linewidth=1.3)
    ax.grid(which='minor', linestyle='-', linewidth='0.7', color='w')
    plt.legend()
    plt.show()


# sorting()
BY, CU, BB = odczytDanych()
data, przychód = określenieDanych(BY, CU, BB)
tupla_lista = listaTupli(data, przychód)
rezultat = przychodyMsc(tupla_lista)
czas, przych = selekcjaDanych(rezultat)
print(wykres(czas, przych))






