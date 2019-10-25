import os
from win32com.client import Dispatch
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.dates import MonthLocator


desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')


def sorting():
    xlApp = Dispatch("Excel.Application")
    xlwb = xlApp.Workbooks.Open(desktop + "\\2014 BAZA MAGRO.xlsx", False, False, None)
    xlAscending = 1
    xlSortColumns = 1
    xlYes = 2
    xlApp.Range("BB5:CU19000").Sort(Key1=xlApp.Range("BB5:CU19000"), Order1=xlAscending,
                                    Header=xlYes, Orientation=xlSortColumns)
    path = ''.join([desktop])
    xlApp.DisplayAlerts = False
    xlwb.SaveAs(path + '\\MODIFY.xlsx')
    xlApp.DisplayAlerts = True
    xlwb.Close()


def odczytDanychBaza():
    start = datetime.now()
    wb = load_workbook(desktop + "\\MODIFY.xlsx", read_only=True, data_only=True)
    ws = wb['BAZA 2014']
    shit = ws['BB5:CU19000']

    daty_baza = []
    przychód_rozl = []

    for BB, bc, bd, be, bf, bg, bh, bi, bj, bk, bl, bm, bn, bo, bp, bq, br, bs, bt, bu, bv, bw, bx, BY, bz, ca, cb, \
        cc, cd, ce, cf, cg, ch, ci, cj, ck, cl, cm, cn, co, cp, cq, cr, cs, ct, CU in shit:

        if BY.value == 'rozl':
            data_raw = str(BB.value)[:7]
            if data_raw != 'None':
                daty_baza.append(data_raw)
                prowizja_rozl = CU.value
                przychód_rozl.append(prowizja_rozl)

    return daty_baza, przychód_rozl


def odczytDanychBank():
    wb = load_workbook("M:/FIRMA MAGRO/BANKI/BZ WBK/HISTORIA/historia - kopia.xlsx", read_only=True, data_only=True)
    ws = wb['CSV srednik (3)']
    shit = ws['A7:H1200']

    daty_bank = []
    rozchód = []
    for A, B, C, D, E, F, G, H in shit:
        if A.value != 'None' and F.value < 0:
            d = A.value[3:]
            daty_bank.append(d)
        if F.value < 0:
            na_plus = F.value * - 1
            rozchód.append(na_plus)

    return daty_bank, rozchód


def daty_przychód(daty_baza, przychód_rozl):
    tpl_lista_baza = []
    for i in zip(daty_baza, przychód_rozl):
        tpl_lista_baza.append(i)

    return tpl_lista_baza


def daty_rozchód(daty_bank, rozchód):
    tpl_lista_bank = []
    for i in zip(daty_bank, rozchód):
        tpl_lista_bank.append(i)

    return tpl_lista_bank


def sumaMscBaza(tpl_lista_baza):
    rezultat = {}
    for k, v in tpl_lista_baza:
        rezultat[k] = rezultat.get(k, 0) + v

    return rezultat


def sumaMscBank(tpl_lista_bank):

    rezultat_bank = {}
    for k, v in tpl_lista_bank:
        rezultat_bank[k] = rezultat_bank.get(k, 0) + v

    return rezultat_bank


def selekcjaDanych(rezultat, rezultat_bank):

    daty_baz = []
    przych_rozl = []
    daty_ban = []
    rozch = []
    for i in rezultat.keys():
        daty_baz.append(i)
    for j in rezultat.values():
        przych_rozl.append(j)
    for k in rezultat_bank.keys():
        daty_ban.append(k)
    for l in rezultat_bank.values():
        rozch.append(l)

    return daty_baz, daty_ban, przych_rozl, rozch


def zysk(przych_rozl, rozch):
    rozch_rev = rozch[::-1]
    zysk = [a - b for a, b in zip(przych_rozl[56:-1], rozch_rev[19:])]

    return zysk


def wykres(daty_baz, przych_rozl, daty_ban, rozch, zysk):

    date_str_baza = []
    date_str_bank = []

    for i in daty_baz:
        if i != 'None':
            date_str_baza.append(datetime.strptime(i, '%Y-%m'))

    for i in daty_ban:
            if i != 'None':
                date_str_bank.append(datetime.strptime(str(i), '%m-%Y'))

    end = datetime.now()
    plt.style.use('ggplot')
    fig, ax = plt.subplots()
    # rotate and align the tick labels so they look better
    fig.autofmt_xdate()
    mloc = MonthLocator()
    ax.xaxis.set_minor_locator(mloc)

    # use a more precise date string for the x axis locations in the toolbar
    ax.fmt_xdata = mdates.DateFormatter('%Y-%m-%d')
    ax.set_title('fig.autofmt_xdate fixes the labels')
    plt.plot(date_str_baza, przych_rozl, color='#001df5', label='zapłacone składki = ' + str('{:.2f}'.format(sum(przych_rozl))))
    plt.plot(date_str_bank, rozch, color='#993344', label='koszty = '
                                + str('{:.2f}'.format(sum(rozch))))  # marker='o'
    plt.plot(date_str_baza[56:-1], zysk, color='#2E8B57', label='zysk = ' + str('{:.2f}'.format(sum(zysk) / 13)))
    # plt.xlabel('rok')
    plt.ylabel('Przychód od 2014 roku w zł')
    plt.title('PRZYCHÓD AGENCJI I KOSZTY SPÓŁKI - 07.10.2019r.')
    # plt.savefig(desktop + "\\przych z bazy.png")
    plt.grid(which='major', color='w', linestyle='-', linewidth=1.3)
    ax.grid(which='minor', linestyle='-', linewidth='0.7', color='w')
    plt.legend()
    plt.show()



# sorting()
daty_baza, przychód_rozl = odczytDanychBaza()
daty_bank, rozchód = odczytDanychBank()
tpl_lista_baza = daty_przychód(daty_baza, przychód_rozl)
tpl_lista_bank = daty_rozchód(daty_bank, rozchód)
rezultat = sumaMscBaza(tpl_lista_baza)
rezultat_bank = sumaMscBank(tpl_lista_bank)
daty_baz, daty_ban, przych_rozl, rozch = selekcjaDanych(rezultat, rezultat_bank)
zysk = zysk(przych_rozl, rozch)
print(wykres(daty_baz, przych_rozl, daty_ban, rozch, zysk))










